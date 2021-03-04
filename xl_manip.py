import openpyxl as xl


def clean_sheet(filename):
    wb = xl.load_workbook(filename)
    ws = wb.active
    min_col = ws.min_column
    max_col = ws.max_column
    min_row = ws.min_row
    max_row = ws.max_row
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            if ws.cell(row, col).value is not None:
                val = str(ws.cell(row, col).value)
                val = val.replace("[", "").replace("]", "").replace("'", "")
                ws.cell(row, col).value = val
    try:
        wb.save(f"{filename}")
    except PermissionError:
        wb.close()
        wb.save(f"{filename}")
    return

def reformat_sheet(filename):
    wb = xl.load_workbook(filename)
    ws = wb.active
    min_col = ws.min_column
    max_col = ws.max_column
    min_row = ws.min_row
    max_row = ws.max_row
    for i in range(max_row):
        min_row = ws.min_row
        val = str(ws.cell(column=3, row=min_row+i).value)
        queue = [i.strip() for i in val.split(",")]
        N = len(queue)
        for j in range(len(queue)):
            ws.cell(column=3 + j, row=min_row+i).value = queue[j]
            for col in range(3, 3+N):
                if col % 2 == 1:
                    ws.cell(row=1, column=col).value = "Departs"
                elif col % 2 == 0:
                    ws.cell(row=1, column=col).value = "Arrives"
    try:
        wb.save(f"{filename}")

    except PermissionError:
        wb.close()
        wb.save(f"{filename}")

    return
